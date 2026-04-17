"""
Apply conditional formatting to a benchmarking CSV or XLSX file.

Colour rules (applied to competitor price cells only):
  Teal  (#008080, white bold text) — competitor price > reference price
                                     (we are cheaper → favourable)
  Red   (#FF0000, white bold text) — competitor price < reference price
                                     (competitor is cheaper → unfavourable)

Reference detection (first match wins):
  "Sales Price (€)"     → reference for all €-denominated competitor columns
  "Sales Price (£)"     → reference for all £-denominated competitor columns
  "DMI Sales Price (€)" → fallback € reference (legacy CSVs without Sales Price)
  "DMI Sales Price (£)" → fallback £ reference (legacy CSVs without Sales Price)

Competitor columns coloured:
  € : DMI Sales Price (€), Dontalia Sales Price (€), Henry Schein Sales Price (€)
  £ : DMI Sales Price (£), DentalSky Sales Price (£)

Usage:
  python format_excel.py                          # uses defaults below
  python format_excel.py input.csv output.xlsx
  python format_excel.py input.xlsx output.xlsx
"""

import csv
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ── defaults ────────────────────────────────────────────────────────────────
DEFAULT_INPUT  = "/home/saj33l/Hashed/Scrapper/output_prices.csv"
DEFAULT_OUTPUT = "/home/saj33l/Hashed/Scrapper/output_prices_formatted.xlsx"

# ── colour styles ────────────────────────────────────────────────────────────
TEAL_FILL  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # light green
RED_FILL   = PatternFill(start_color="FFB6B6", end_color="FFB6B6", fill_type="solid")  # light red
WHITE_FONT = Font(color="000000")

# ── currency buckets ─────────────────────────────────────────────────────────
# Each entry: (reference column candidates in priority order, competitor columns)
CURRENCY_GROUPS = [
    {
        "refs": ["Sales Price (€)", "DMI Sales Price (€)"],
        "competitors": [
            "DMI Sales Price (€)",
            "Dontalia Sales Price (€)",
            "Henry Schein Sales Price (€)",
        ],
    },
    {
        "refs": ["Sales Price (£)", "DMI Sales Price (£)"],
        "competitors": [
            "DMI Sales Price (£)",
            "DentalSky Sales Price (£)",
        ],
    },
]


def _to_float(val):
    if not val or str(val).strip() in ("N/A", "-", "Not listed on competitor", "nan", ""):
        return None
    m = re.search(r"(\d[\d,]*\.?\d*)", str(val).replace(",", ""))
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _load_rows_from_csv(path: str):
    with open(path, encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    return reader[0], reader[1:]


def _load_rows_from_xlsx(path: str):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [[str(c) if c is not None else "" for c in r] for r in rows[1:]]
    return headers, data


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _resolve_comparisons(headers):
    """
    Return a list of (ref_col_idx, comp_col_idx) pairs to format,
    based on the columns present in headers.
    """
    col_map = {h: i for i, h in enumerate(headers)}
    pairs = []

    for group in CURRENCY_GROUPS:
        # Pick the first available reference column
        ref_idx = None
        ref_name = None
        for candidate in group["refs"]:
            if candidate in col_map:
                ref_idx = col_map[candidate]
                ref_name = candidate
                break
        if ref_idx is None:
            continue

        # Collect competitor columns (skip if it IS the reference column)
        for comp_name in group["competitors"]:
            if comp_name not in col_map:
                continue
            c_idx = col_map[comp_name]
            if c_idx == ref_idx:          # don't compare a column against itself
                continue
            pairs.append((ref_idx, c_idx))

    return pairs


def apply_formatting(headers, data_rows, ws):
    """Write headers + data into ws, then apply teal/red formatting."""
    ws.append(headers)

    comparisons = _resolve_comparisons(headers)

    for r_idx, row in enumerate(data_rows, start=2):
        ws.append(list(row))
        for ref_idx, comp_idx in comparisons:
            ref_val  = _to_float(row[ref_idx]  if ref_idx  < len(row) else None)
            comp_val = _to_float(row[comp_idx] if comp_idx < len(row) else None)
            if ref_val is None or comp_val is None:
                continue
            cell = ws.cell(row=r_idx, column=comp_idx + 1)
            if comp_val > ref_val:
                cell.fill = TEAL_FILL
                cell.font = WHITE_FONT
            elif comp_val < ref_val:
                cell.fill = RED_FILL
                cell.font = WHITE_FONT


def main():
    input_file  = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if input_file.lower().endswith((".xlsx", ".xls")):
        headers, data_rows = _load_rows_from_xlsx(input_file)
    else:
        headers, data_rows = _load_rows_from_csv(input_file)

    if not headers:
        print("Empty input file — nothing to do.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmarking Data"

    apply_formatting(headers, data_rows, ws)
    _autofit(ws)

    wb.save(output_file)
    print(f"Saved formatted Excel → {output_file}")


if __name__ == "__main__":
    main()

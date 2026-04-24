"""
Push benchmarking CSV/XLSX data to Google Sheets with conditional colour formatting.

Colour rules (same logic as format_excel.py):
  Light green (#90EE90) — competitor price > our price  (we are cheaper)
  Light red   (#FFB6B6) — competitor price < our price  (competitor is cheaper)

Usage:
  python push_to_sheets.py                          # uses DEFAULT_INPUT, DEFAULT_SHEET_ID
  python push_to_sheets.py input.csv SHEET_ID
  python push_to_sheets.py input.xlsx SHEET_ID
"""

import csv
import re
import sys
import os

import gspread
from google.oauth2.service_account import Credentials

# ── defaults ─────────────────────────────────────────────────────────────────
DEFAULT_INPUT    = "output_prices.csv"
DEFAULT_SHEET_ID = "1qJCFzTea15Q9HlWqzIEFKHn47PW-Nak-CUHNulZtsjs"
CREDS_FILE       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bfm-competitor-price-scraper-60bbef18550e.json")
WORKSHEET_NAME   = "Benchmarking Data"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ── colour hex values ─────────────────────────────────────────────────────────
GREEN_HEX = "90EE90"   # competitor more expensive (we are cheaper)
RED_HEX   = "FFB6B6"   # competitor cheaper (unfavourable)

# ── currency comparison groups (mirrors format_excel.py) ─────────────────────
CURRENCY_GROUPS = [
    {
        "refs": ["Sales Price (€)", "DMI Sales Price (€)"],
        "competitors": [
            "DMI Sales Price (€)",
            "Dontalia Sales Price (€)",
            "Henry Schein Sales Price (€)",
        ],
    },
    {
        "refs": ["Sales Price (£)", "DMI Sales Price (£)"],
        "competitors": [
            "DMI Sales Price (£)",
            "DentalSky Sales Price (£)",
        ],
    },
]


def _to_float(val):
    if not val or str(val).strip() in ("N/A", "-", "Not listed on competitor", "nan", ""):
        return None
    m = re.search(r"(\d[\d,]*\.?\d*)", str(val).replace(",", ""))
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _load_csv(path):
    with open(path, encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    return reader[0], reader[1:]


def _load_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Prices"] if "Prices" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [[str(c) if c is not None else "" for c in r] for r in rows[1:]]
    return headers, data


def _resolve_comparisons(headers):
    col_map = {h: i for i, h in enumerate(headers)}
    pairs = []
    for group in CURRENCY_GROUPS:
        ref_idx = None
        for candidate in group["refs"]:
            if candidate in col_map:
                ref_idx = col_map[candidate]
                break
        if ref_idx is None:
            continue
        for comp_name in group["competitors"]:
            if comp_name not in col_map:
                continue
            c_idx = col_map[comp_name]
            if c_idx == ref_idx:
                continue
            pairs.append((ref_idx, c_idx))
    return pairs


def _col_letter(n):
    """Convert 0-based column index to A1-notation letter(s)."""
    result = ""
    n += 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def build_format_requests(sheet_id, headers, data_rows, comparisons):
    """Build a list of Sheets API batchUpdate requests for cell colours."""
    requests = []

    for row_idx, row in enumerate(data_rows):
        gsheet_row = row_idx + 1  # 0-based data + 1 header row = 1-based row

        for ref_idx, comp_idx in comparisons:
            ref_val  = _to_float(row[ref_idx]  if ref_idx  < len(row) else None)
            comp_val = _to_float(row[comp_idx] if comp_idx < len(row) else None)
            if ref_val is None or comp_val is None:
                continue

            if comp_val > ref_val:
                hex_color = GREEN_HEX
            elif comp_val < ref_val:
                hex_color = RED_HEX
            else:
                continue

            r = int(hex_color[0:2], 16) / 255
            g = int(hex_color[2:4], 16) / 255
            b = int(hex_color[4:6], 16) / 255

            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": gsheet_row,
                        "endRowIndex": gsheet_row + 1,
                        "startColumnIndex": comp_idx,
                        "endColumnIndex": comp_idx + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": r, "green": g, "blue": b}
                        }
                    },
                    "fields": "userEnteredFormat.backgroundColor",
                }
            })

    return requests


def push(input_file, spreadsheet_id, worksheet_name=WORKSHEET_NAME):
    print(f"Loading data from: {input_file}")
    if input_file.lower().endswith((".xlsx", ".xls")):
        headers, data_rows = _load_xlsx(input_file)
    else:
        headers, data_rows = _load_csv(input_file)

    print(f"Rows: {len(data_rows)}  Columns: {len(headers)}")

    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
    gc = gspread.authorize(creds)

    spreadsheet = gc.open_by_key(spreadsheet_id)

    # Get or create the worksheet
    try:
        ws = spreadsheet.worksheet(worksheet_name)
        print(f"Found existing worksheet: '{worksheet_name}' — clearing values and formatting")
        ws.clear()
        # clear() only removes values — explicitly reset all cell formatting too
        spreadsheet.batch_update({"requests": [{
            "repeatCell": {
                "range": {"sheetId": ws.id},
                "cell": {"userEnteredFormat": {}},
                "fields": "userEnteredFormat",
            }
        }]})
    except gspread.exceptions.WorksheetNotFound:
        print(f"Creating worksheet: '{worksheet_name}'")
        ws = spreadsheet.add_worksheet(
            title=worksheet_name,
            rows=max(len(data_rows) + 10, 1000),
            cols=max(len(headers) + 5, 50),
        )

    # Upload all data at once
    all_rows = [headers] + [list(row) for row in data_rows]
    print(f"Uploading {len(all_rows)} rows to Google Sheets...")
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print("Data upload complete.")

    # Bold the header row
    header_fmt = {
        "repeatCell": {
            "range": {
                "sheetId": ws.id,
                "startRowIndex": 0,
                "endRowIndex": 1,
            },
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {"bold": True},
                    "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.9},
                }
            },
            "fields": "userEnteredFormat(textFormat,backgroundColor)",
        }
    }

    # Freeze header row
    freeze_fmt = {
        "updateSheetProperties": {
            "properties": {
                "sheetId": ws.id,
                "gridProperties": {"frozenRowCount": 1},
            },
            "fields": "gridProperties.frozenRowCount",
        }
    }

    # Build colour formatting requests
    comparisons = _resolve_comparisons(headers)
    colour_requests = build_format_requests(ws.id, headers, data_rows, comparisons)

    all_requests = [header_fmt, freeze_fmt] + colour_requests

    print(f"Applying formatting ({len(colour_requests)} colour cells)...")
    spreadsheet.batch_update({"requests": all_requests})
    print("Formatting applied.")

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    print(f"\nDone! View sheet: {url}")


def main():
    input_file     = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    spreadsheet_id = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SHEET_ID
    push(input_file, spreadsheet_id)


if __name__ == "__main__":
    main()
